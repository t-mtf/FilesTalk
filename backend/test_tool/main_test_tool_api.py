import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Literal, Optional

import numpy as np
import pandas as pd
from dotenv import find_dotenv, load_dotenv
from langchain_openai import AzureChatOpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pydantic import BaseModel
from sklearn.metrics import (
    accuracy_score,
    confusion_matrix,
    precision_score,
    recall_score,
)
import os


def configure_logging(
    logger_name: str,
    filename: str = "app.log",
    level: int = logging.ERROR,
    log_format: str = "%(asctime)s:%(levelname)s:%(message)s",
) -> None:
    """
    Configure logging settings for the application.

    This function sets up the logging system to log messages to a file.
    Args:
        logger_name (str): The name of the logger.
        filename (str): The name of the file to log messages to.
        level (int): The minimum severity level for messages to be logged.
        log_format (str): The format of log messages.

    Returns:
        None
    """
    log_file_path = Path(".") / "data" / filename
    log_file_path.parent.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger(logger_name)
    handler = logging.FileHandler(log_file_path)
    formatter = logging.Formatter(log_format)
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.setLevel(level)


def configure_azure_settings(
    chat_model: str = "gpt-4-128k", embedding_model: str = "text-embedding-ada-002"
) -> dict:
    """
    Configure and return Azure settings for various services like
    Azure OpenAI.

    Args:
        chat_model (str): The chat model to use.
        embedding_model (str): The embedding model to use.

    Returns:
        dict: A dictionary containing the configuration settings for Azure services.
    """
    batch_logger = logging.getLogger("batch_logger")

    azure_settings = {}

    # Fetch environment variables based on region

    azure_settings["azure_openai_endpoint"] = os.getenv("AZURE_OPENAI_ENDPOINT")
    azure_key = os.getenv("AZURE_OPENAI_API_KEY")
    if azure_key is None:
        batch_logger.error("Environment variable 'AZURE_OPENAI_API_KEY' is not set.")
        raise ValueError("Environment variable 'AZURE_OPENAI_API_KEY' is not set.")

    azure_settings["azure_openai_key"] = azure_key
    if embedding_model == "text-embedding-ada-002":

        azure_settings["azure_openai_embedding_deployment"] = os.getenv(
            "TEXT_EMBEDDING_ADA_MODEL_DEPLOYMENT"
        )
        azure_settings["azure_openai_embedding_api_version"] = os.getenv(
            "TEXT_EMBEDDING_ADA_API_VERSION"
        )
    elif embedding_model == "text-embedding-3-large":

        azure_settings["azure_openai_embedding_deployment"] = os.getenv(
            "TEXT_EMBEDDING_LARGE_MODEL_DEPLOYMENT"
        )
        azure_settings["azure_openai_embedding_api_version"] = os.getenv(
            "TEXT_EMBEDDING_LARGE_API_VERSION"
        )

    if chat_model == "gpt-4-128k":
        azure_settings["azure_openai_chat_deployment"] = os.getenv(
            "GPT4_128K_MODEL_DEPLOYMENT"
        )
        azure_settings["azure_openai_chat_api_version"] = os.getenv(
            "GPT4_128K_AZURE_API_VERSION"
        )

        azure_settings["azure_openai_chat_model_name"] = chat_model

    elif chat_model == "gpt-4o":

        azure_settings["azure_openai_chat_deployment"] = os.getenv(
            "GPT4O_MODEL_DEPLOYMENT"
        )
        azure_settings["azure_openai_chat_api_version"] = os.getenv(
            "GPT4O_AZURE_API_VERSION"
        )

    return azure_settings


CHAT_MODEL = "gpt-4o"
EMBEDDING_MODEL = "text-embedding-3-large"
TEMPERATURE = 0.1


class Comments(BaseModel):
    comments: Optional[str]


class InflationEffectiveDate(BaseModel):
    inflation_effective_date: Optional[str]


class InflationType(BaseModel):
    inflation_type: Optional[List[Literal["General", "3rd Party", "Labour", "OEM"]]]


class InflationNoticePeriod(BaseModel):
    inflation_notice_period: Optional[Literal["0", "30", "60", "90"]]


class Periodicity(BaseModel):
    periodicity: Optional[Literal["Annually"]]


class InflationFound(BaseModel):
    inflation_found: Literal[True, False]


class ResponseSimilarity(BaseModel):
    similarity: Literal["identical", "nearly identical", "distinct"]


column_to_class_map = {
    "inflation_found": InflationFound,
    "inflation_type": InflationType,
    "inflation_effective_date": InflationEffectiveDate,
    "inflation_notice_period": InflationNoticePeriod,
    "periodicity": Periodicity,
    "comments": Comments,
    "world_bank": Comments,
    "language": Comments,
}


def synthesize_function(llm: AzureChatOpenAI, context: str) -> dict:
    """
    This function synthesizes all the results corresponding to a contract number.
    Parameters:
        llm (AzureChatOpenAI): The model used for the comparison.
        context (str): All responses for a contract_number

    Returns:
        dict: The synthesized response as a dictionary.
    """

    try:
        logger = logging.getLogger("logger")
        qa_system_prompt = f"""
    As an AI assistant specializing in contract analysis, you will be given the necessary context. 

    Your task is to generate an answer based solely on the provided context. This context includes a 
    question and several responses obtained from different documents. 
    Your goal is to synthesize the responses. If the answer is not provided in the context return None. 
    **Context:**

    {context}
    """
        human_prompt = """
    Based on the provided context, synthesize the responses. Return a JSON object with the provided output format.

    
    Here are some examples:

    Find the answer to this question from the provided responses: How many days the customer needs to be noticed before the price adjustment can take place ?
    Responses in different files:
    file_name: file1.docx
    inflation_notice_period: nan
    file_name: file2.pdf
    inflation_notice_period: The customer needs to be given at least ninety (90) days' prior written notice before the price adjustment can take place.

    Expected output:
    {"inflation_notice_period":"90"}


    Find the answer to this question from the provided responses: How many days the customer needs to be noticed before the price adjustment can take place ?
    Responses in different files:
    file_name: file4.docx
    inflation_notice_period: nan
    file_name: file5.pdf
    inflation_notice_period: nan

    Expected output:
    {"inflation_notice_period": None}

    """

        messages = [
            ("system", qa_system_prompt),
            ("human", human_prompt),
        ]

        response = llm.invoke(
            messages,
        )

        return response
    except json.JSONDecodeError as e:
        logger.error(e)
        return response.content


def format_and_save_excel(df, output_file_path, sheet_name):
    try:
        logger = logging.getLogger("logger")
        writer = pd.ExcelWriter(output_file_path, engine="xlsxwriter")
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        format_blue = workbook.add_format(
            {"bg_color": "#C5D9F1", "font_color": "#000000", "bold": True}
        )
        format_green = workbook.add_format(
            {"bg_color": "#9BBB59", "font_color": "#000000", "bold": True}
        )

        for col_num, column_name in enumerate(df.columns):
            if column_name in ["customer_name", "contract_number"]:
                worksheet.write(0, col_num, column_name, format_blue)
            else:
                worksheet.write(0, col_num, column_name, format_green)

        writer.close()

    except Exception as e:
        logger.error(e)


def main_synthesize(input_file_name):

    try:

        logger = logging.getLogger("logger")

        load_dotenv(find_dotenv())
        azure_settings = configure_azure_settings(
            chat_model=CHAT_MODEL, embedding_model=EMBEDDING_MODEL
        )

        llm = AzureChatOpenAI(
            openai_api_key=azure_settings["azure_openai_key"],
            azure_endpoint=azure_settings["azure_openai_endpoint"],
            openai_api_version=azure_settings["azure_openai_chat_api_version"],
            azure_deployment=azure_settings["azure_openai_chat_deployment"],
            model=CHAT_MODEL,
            temperature=TEMPERATURE,
            timeout=120,
            max_retries=3,
            seed=123,
        )

        input_file_path = Path(".") / "data" / "test_dataset.xlsx"
        df0 = pd.read_excel(
            input_file_path,
            sheet_name="Sheet1",
            dtype=str,
        )

        input_file_path = Path(".") / "data" / f"{input_file_name}.xlsx"

        df = pd.read_excel(
            input_file_path,
            sheet_name="Results",
            dtype=str,
        )

        df = df[df["contract_number"].isin(df0["contract_number"])]
        # df = df[df["is_new"] == "YES"]

        df = df.rename(
            columns={
                "inflation found": "inflation_found",
                "inflation type": "inflation_type",
                "inflation effective date": "inflation_effective_date",
                "inflation notice period": "inflation_notice_period",
                "periodicity": "periodicity",
                "customerName": "customer_name",
                "original_file_name": "filename",
                "world bank": "world_bank",
                # "Comments": "comments",
            }
        )
        results = []

        selected_columns = [
            "customer_name",
            "inflation_found",
            "inflation_type",
            "inflation_effective_date",
            "inflation_notice_period",
            "periodicity",
            "world_bank",
            "language",
        ]
        #            "comments",

        prompts = [
            {"name": "inflation_found", "value": "Is any inflation clause found?"},
            {
                "name": "inflation_type",
                "value": """What is inflation type? Inflation can be categorized into different types based on its source and impact. Below are the types of inflation you may encounter:
        General: This refers to the overall increase in prices across the economy, and it is typically measured by indices such as the Consumer Price Index (CPI).
        3rd Party: This type of inflation occurs when external suppliers or service providers increase their prices.
        Labour: Labour inflation is characterized by rising wages and benefits costs.
        OEM: Original Equipment Manufacturer (OEM) inflation occurs when manufacturers of components or products increase their prices.""",
            },
            {
                "name": "inflation_effective_date",
                "value": "What is the inflation effective date? Please provide the answer in this format: 'DD/MM/YYYY', e.g., '20/06/2024'.",
            },
            {
                "name": "inflation_notice_period",
                "value": "How many days the customer needs to be noticed before the price adjustment can take place?",
            },
            {
                "name": "periodicity",
                "value": "What is the periodicity of the price adjustment due to the inflation clause? Examples: Monthly, annually, etc. If periodicity does not exist in the document answer NA.",
            },
        ]

        # Context:

        #         {
        #     "name": "comments",
        #     "value": "What are the section numbers of inflation clauses?",
        # },

        for contract_number in df["contract_number"].unique():

            df_filtered = df[df["contract_number"] == contract_number]

            contract_result = {"contract_number": contract_number}

            for column in selected_columns:

                if column == "customer_name":
                    contract_result[column] = df_filtered[column].iloc[0]
                    continue
                prompt = next(
                    (item for item in prompts if item["name"] == column), None
                )
                if prompt is not None:
                    prompt_value = prompt["value"]
                else:
                    prompt_value = "Prompt not found"

                context = f"Find the answer to this question from the provided responses: {prompt_value}"

                for _, row in df_filtered.iterrows():

                    context += f"\nResponses in different files:\nfile_name: {row['filename']}\n"
                    context += f"{column}: {row[column]}\n"

                logger.info(context)

                model_class = column_to_class_map.get(column)
                if model_class:
                    response_llm_structured = synthesize_function(
                        llm.with_structured_output(model_class), context
                    )

                    attribute_value = getattr(response_llm_structured, column, None)

                    if isinstance(attribute_value, list):
                        contract_result[column] = ", ".join(attribute_value)
                    else:
                        contract_result[column] = attribute_value

            results.append(contract_result)

        df_synthesized = pd.DataFrame(results)
        output_file_path = (
            Path(".")
            / "data"
            / f"output-{input_file_name}-{datetime.now().strftime("%Y-%m-%d-%H-%M-%S")}.xlsx"
        )
        # df_synthesized.to_excel(
        #     output_file_path,
        #     index=False,
        # )
        format_and_save_excel(
            df=df_synthesized, output_file_path=output_file_path, sheet_name="Results"
        )
        return output_file_path
    except Exception as e:
        logger.error("An error occured in main function: %s", e)


def compare_function(llm: AzureChatOpenAI, result1: str, result2: str) -> str:
    """
    This function compares two text results and determines their similarity level
    using a large language model as "identical", "nearly identical" or "distinct".

    Parameters:
        llm (AzureChatOpenAI): The model used for the comparison.
        result1 (str): The first text to be compared.
        result2 (str): The second text to be compared.

    Returns:
        str: The similarity level of the two texts.
    """

    try:
        logger = logging.getLogger("logger")
        if (pd.isna(result1) or result1.strip() == "") and (
            pd.isna(result2) or result2.strip() == ""
        ):
            return ResponseSimilarity(similarity="identical")

        context = f"Text1:\n{result1} \n Text2:\n{result2}"

        qa_system_prompt = f"""
    You are an AI assistant specialized in contract analysis. You will be \
    provided with the necessary context.

    Please generate an answer based solely on the provided context. Ensure that your response is in English.

    **Context:**

    {context}
    """
        human_prompt = """
    Please compare the provided texts in the context and determine whether the meaning is identical, nearly identical
    or distinct. If the exact texts vary but still convey the same message, it should be considered as identical. If one
    text is empty while the other one has some value it should be considered as "distinct".
    Return a JSON object with key "similarity" and a value of "identical", "nearly identical" or "distinct"

    Here are some examples:

    Text1:
    This contract shall be governed by and interpreted in accordance with the laws of France. Any disputes arising from this contract will be resolved in the courts of France.

    Text 2:
    This contract will be governed by and interpreted following the laws of France. Any disputes that arise from this contract shall be settled in the courts of France.

    Expected output:
    {"similarity":"identical"}
    ====================================
    Text1:
    World Bank

    **Clause 1.4**: "The inflation rate will be based on the most recent annual headline consumer price index (CPI) inflation information (series name “Headline Consumer Price Inflation”) for the United States, as published by the **World Bank** at https://www.worldbank.org/en/research/brief/inflation-database (the “Index”)."

    Text 2:
    World Bank, United States

    Expected output:
    {"similarity":"identical"}
    ====================================
    Text1:
    "annually

    **Clause 5.6 Charges Adjustment**: upon each anniversary of the Effective Date (the “Revision Date”), Orange reserves the right to adjust its Charges by reference to the most recent annual headline consumer price index (CPI) inflation information...

    Text 2:
    Annually

    Expected output:
    {"similarity":"identical"}
    ====================================
    Text1:
    "Annually"

    Text 2:
    ""

    Expected output:
    {"similarity":"distinct"}

    """

        messages = [
            ("system", qa_system_prompt),
            ("human", human_prompt),
        ]

        response = llm.invoke(
            messages,
        )

        return response
    except Exception as e:
        logger.error(e)
        return ResponseSimilarity(similarity="distinct")


def main_compare(
    reference_file_name: str,
    reference_suffix: str,
    input_file_name: str,
    input_suffix: str,
) -> Path:
    """
    This function compares the results of inflation analysis between an input file and a reference file.

    Parameters:
        reference_file_name (str): The name of the reference file for comparison.
        reference_suffix (str): The suffix to be added to columns from the reference file.
        input_file_name (str): The name of the input file.
        input_suffix (str): The suffix to be added to columns from the input file.

    Returns:
        Path: The path to the Excel file containing the comparison results.
    """
    try:

        logger = logging.getLogger("logger")

        load_dotenv(find_dotenv())
        azure_settings = configure_azure_settings(
            chat_model=CHAT_MODEL, embedding_model=EMBEDDING_MODEL
        )

        llm = AzureChatOpenAI(
            openai_api_key=azure_settings["azure_openai_key"],
            azure_endpoint=azure_settings["azure_openai_endpoint"],
            openai_api_version=azure_settings["azure_openai_chat_api_version"],
            azure_deployment=azure_settings["azure_openai_chat_deployment"],
            model=CHAT_MODEL,
            temperature=TEMPERATURE,
            timeout=120,
            max_retries=3,
            seed=123,
        )
        structured_llm = llm.with_structured_output(ResponseSimilarity)

        input_file_path = Path(".") / "data" / f"{input_file_name}.xlsx"
        reference_file_path = Path(".") / "data" / f"{reference_file_name}.xlsx"

        df1 = pd.read_excel(
            reference_file_path, sheet_name="Sheet1", dtype=str, na_values=[""]
        )
        df2 = pd.read_excel(
            input_file_path, sheet_name="Results", dtype=str, na_values=[""]
        )

        df_merged = pd.merge(
            df1,
            df2,
            on=["contract_number", "customer_name"],
            suffixes=(f"_{reference_suffix}", f"_{input_suffix}"),
        )

        columns_to_compare = [
            "inflation_found",
            "inflation_type",
            "inflation_effective_date",
            "inflation_notice_period",
            "periodicity",
        ]
        columns_to_keep = ["contract_number", "customer_name"]

        for column in columns_to_compare:
            df_merged[f"{column}_comparison"] = df_merged.apply(
                lambda row: (
                    compare_result := compare_function(
                        structured_llm,
                        row[f"{column}_{reference_suffix}"],
                        row[f"{column}_{input_suffix}"],
                    ),
                    compare_result.similarity if compare_result is not None else np.nan,
                )[1],
                axis=1,
            )
        for column in columns_to_compare:
            columns_to_keep.extend(
                [
                    f"{column}_{reference_suffix}",
                    f"{column}_{input_suffix}",
                    f"{column}_comparison",
                ]
            )

        df_merged = df_merged[columns_to_keep]

        output_excel_file_path = (
            Path(".")
            / "data"
            / f"COMPARISON_{input_suffix}_{reference_suffix}_results.xlsx"
        )
        df_merged.to_excel(output_excel_file_path, sheet_name="Results", index=False)

        format_comparison_results(
            input_file_path=output_excel_file_path,
            sheet_name="Results",
        )
        return output_excel_file_path
    except Exception as e:
        logger.error(e)


def format_comparison_results(input_file_path, sheet_name="Results"):
    """This function highlights comparison results based on their value"""
    try:
        logger = logging.getLogger("logger")
        output_file_path = Path(".") / "data" / f"{input_file_path.stem}_formatted.xlsx"

        workbook = load_workbook(input_file_path)
        worksheet = workbook[sheet_name]

        df = pd.read_excel(input_file_path, sheet_name=sheet_name)

        comparison_columns = [col for col in df.columns if col.endswith("_comparison")]

        fill_identical = PatternFill(
            start_color="00FF00", end_color="00FF00", fill_type="solid"
        )
        fill_nearly_identical = PatternFill(
            start_color="FFA500", end_color="FFA500", fill_type="solid"
        )
        fill_distinct = PatternFill(
            start_color="FFC0CB", end_color="FFC0CB", fill_type="solid"
        )

        for col in comparison_columns:
            col_idx = df.columns.get_loc(col) + 1
            for row in worksheet.iter_rows(
                min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value == "identical":
                        cell.fill = fill_identical
                    elif cell.value == "nearly identical":
                        cell.fill = fill_nearly_identical
                    elif cell.value == "distinct":
                        cell.fill = fill_distinct

        workbook.save(output_file_path)
    except Exception as e:
        logger.error(e)


def calculate_metrics(
    input_file_path: Path, extraction_version: str, output_file_path: Path
):
    """
    This function evaluates inflation clause-related information extracted from contract documents.
    It calculates performance metrics based on the comparison results between the extracted data and
    the reference dataset, and formats the original data sheet with colored cells based on comparison results.

    Parameters:
        input_file_path (Path): Path to the Excel file containing comparison results.
        extraction_version (str): The version identifier for the extraction method used.
        output_file_path (Path): Path to save the Excel file with calculated metrics.

    Returns:
        None: The function writes results directly to the specified output file.
    """

    try:
        logger = logging.getLogger("logger")
        df = pd.read_excel(
            input_file_path, sheet_name="Results", dtype=str, na_values=[""]
        )

        y_true = df["inflation_found_reference"].map({"True": True, "False": False})
        y_pred = df[f"inflation_found_{extraction_version}"].map(
            {"True": True, "False": False}
        )

        tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()

        accuracy = accuracy_score(y_true, y_pred)
        precision = precision_score(y_true, y_pred)
        recall = recall_score(y_true, y_pred)

        metrics_data = {
            "Metric": [
                "True Positives",
                "True Negatives",
                "False Positives",
                "False Negatives",
                "Accuracy",
                "Precision",
                "Recall",
            ],
            "Value": [tp, tn, fp, fn, accuracy, precision, recall],
        }
        metrics_df = pd.DataFrame(metrics_data)

        comparison_metrics = []
        for column in df.columns:
            if (
                column.endswith("_comparison")
                and column != "inflation_found_comparison"
            ):
                total_identical = df[column].str.count("identical").sum()
                total_distinct = df[column].str.count("distinct").sum()
                total = total_identical + total_distinct

                accuracy = total_identical / total if total > 0 else 0
                comparison_metrics.append(
                    {
                        "Comparison Metric": column.replace("_comparison", ""),
                        "Total Identical": total_identical,
                        "Total Distinct": total_distinct,
                        "Accuracy": accuracy,
                    }
                )

        comparison_metrics_df = pd.DataFrame(comparison_metrics)

        # Create an Excel writer and write the dataframes to it
        with pd.ExcelWriter(output_file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Original Data", index=False)
            metrics_df.to_excel(
                writer, sheet_name="Clause Identification Metrics", index=False
            )
            comparison_metrics_df.to_excel(
                writer, sheet_name="Accuracy for each field", index=False
            )

            # Access the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = workbook["Original Data"]

            # Define the fills for formatting
            fill_identical = PatternFill(
                start_color="00FF00", end_color="00FF00", fill_type="solid"
            )
            fill_nearly_identical = PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            )
            fill_distinct = PatternFill(
                start_color="FFC0CB", end_color="FFC0CB", fill_type="solid"
            )

            # Apply formatting based on comparison results
            comparison_columns = [
                col for col in df.columns if col.endswith("_comparison")
            ]
            for col in comparison_columns:
                col_idx = df.columns.get_loc(col) + 1
                for row in worksheet.iter_rows(
                    min_row=2,
                    max_row=worksheet.max_row,
                    min_col=col_idx,
                    max_col=col_idx,
                ):
                    for cell in row:
                        if cell.value == "identical":
                            cell.fill = fill_identical
                        elif cell.value == "nearly identical":
                            cell.fill = fill_nearly_identical
                        elif cell.value == "distinct":
                            cell.fill = fill_distinct

    except Exception as e:
        logger.error(e)


def compare_two_extractions(
    input_file_name1, file_suffix1, input_file_name2, file_suffix2
):

    input_file_name1 = "document_IZ_contracts_fields_dec_2024_v4"
    input_file_name2 = "document_IZ_contracts_fields_dec_2024_v5"
    file_suffix1 = "v4"
    file_suffix2 = "v5"
    output_file_path_synthesis1 = main_synthesize(input_file_name1)
    output_file_path_synthesis2 = main_synthesize(input_file_name2)

    main_compare(
        str(output_file_path_synthesis1.stem),
        file_suffix1,
        str(output_file_path_synthesis2.stem),
        file_suffix2,
    )


def main():
    """This function compares the results of an extraction for inflation clause
    with a reference test dataset and synthesizes the results and finally calculates
    some performance metrics. Place the files in "data" directory.
    It can also compare two extractions.
    """
    try:
        logger = logging.getLogger("logger")

        compare_two_results = False
        logger.info("compare_two_results: %s", compare_two_results)
        if compare_two_results:
            logger.info("compare_two_results block is being executed.")
            input_file_name1 = "document_IZ_contracts_fields_dec_2024_v4"
            file_suffix1 = "v4"
            input_file_name2 = "document_IZ_contracts_fields_dec_2024_v5"
            file_suffix2 = "v5"
            compare_two_extractions(
                input_file_name1, file_suffix1, input_file_name2, file_suffix2
            )
        else:

            reference_file_name = "test_dataset"
            reference_suffix = "reference"
            input_file_name = (
                "00de8f4e-d746-442c-9267-8e1f32cf0f3e"  # extension .xlsx is not needed.
            )
            input_suffix = f"api-00d"

            output_file_path_synthesis = main_synthesize(input_file_name)

            comparison_file_path = main_compare(
                reference_file_name,
                reference_suffix,
                str(output_file_path_synthesis.stem),
                input_suffix,
            )

            output_file_path_metrics = (
                Path(".") / "data" / f"comparison_{input_suffix}_reference_metrics.xlsx"
            )
            calculate_metrics(
                comparison_file_path, input_suffix, output_file_path_metrics
            )
    except Exception as e:
        logger.error(e)


if __name__ == "__main__":
    configure_logging(
        logger_name="logger",
        filename=f"{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.log",
        level=logging.INFO,
    )
    main()
